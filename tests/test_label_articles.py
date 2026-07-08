"""
Unit tests for tools/label_articles_by_outlet.py.

Covers:
  - Domain normalisation (_normalise_domain)
  - AllSides index building (_build_allsides_index)
  - Outlet lookup (_lookup): exact, subdomain, prefix, no-match
  - Full pipeline (run): CSV, Excel, labeled folders, dry-run, error paths
"""

import csv
import json
import pytest
from pathlib import Path

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from tools.label_articles_by_outlet import (
    _normalise_domain,
    _build_allsides_index,
    _lookup,
    run,
    LABEL_FOLDERS,
)

from tests.conftest import ALLSIDES_CSV, EXPECTED_LABELS


# ---------------------------------------------------------------------------
# _normalise_domain
# ---------------------------------------------------------------------------

class TestNormaliseDomain:
    def test_bare_domain(self):
        assert _normalise_domain("cnn.com") == "cnn.com"

    def test_full_url_with_www_and_slash(self):
        assert _normalise_domain("https://www.cnn.com/") == "cnn.com"

    def test_http_scheme(self):
        assert _normalise_domain("http://foxnews.com") == "foxnews.com"

    def test_uppercase_stripped(self):
        assert _normalise_domain("HTTPS://WWW.CNN.COM/") == "cnn.com"

    def test_port_stripped(self):
        assert _normalise_domain("https://example.com:8080/path") == "example.com"

    def test_subdomain_preserved(self):
        # Subdomains are kept intact - matching logic handles them separately
        assert _normalise_domain("politics.nytimes.com") == "politics.nytimes.com"

    def test_no_www_prefix(self):
        assert _normalise_domain("www.bbc.co.uk") == "bbc.co.uk"

    def test_trailing_whitespace(self):
        assert _normalise_domain("  nytimes.com  ") == "nytimes.com"

    def test_path_ignored(self):
        assert _normalise_domain("https://www.reuters.com/world/") == "reuters.com"


# ---------------------------------------------------------------------------
# _build_allsides_index
# ---------------------------------------------------------------------------

class TestBuildAllsidesIndex:
    def test_loads_real_csv(self, allsides_csv_path):
        index = _build_allsides_index(allsides_csv_path)
        assert len(index) > 1000, "Expected at least 1000 outlets in AllSides CSV"

    def test_known_outlet_present(self, allsides_csv_path):
        index = _build_allsides_index(allsides_csv_path)
        # CNN is in AllSides
        assert "cnn.com" in index or any("cnn" in k for k in index)

    def test_values_are_name_label_tuples(self, allsides_csv_path):
        index = _build_allsides_index(allsides_csv_path)
        for domain, value in list(index.items())[:10]:
            assert isinstance(value, tuple) and len(value) == 2
            name, label = value
            assert isinstance(name, str)
            assert label in {"Left", "Lean Left", "Center", "Lean Right", "Right", "Mixed"}

    def test_skips_rows_with_empty_url(self, tmp_path):
        csv_path = tmp_path / "test.csv"
        csv_path.write_text(
            "allsides_media_bias_ratings/publication/source_name,"
            "allsides_media_bias_ratings/publication/source_type,"
            "allsides_media_bias_ratings/publication/media_bias_rating,"
            "allsides_media_bias_ratings/publication/source_url,"
            "allsides_media_bias_ratings/publication/allsides_url\n"
            "Good Source,News Media,Center,https://www.example.com/,https://allsides.com/1\n"
            "Bad Source,News Media,Center,,https://allsides.com/2\n",
            encoding="utf-8",
        )
        index = _build_allsides_index(csv_path)
        assert "example.com" in index
        assert len(index) == 1  # empty URL row skipped

    def test_skips_rows_with_empty_label(self, tmp_path):
        csv_path = tmp_path / "test.csv"
        csv_path.write_text(
            "allsides_media_bias_ratings/publication/source_name,"
            "allsides_media_bias_ratings/publication/source_type,"
            "allsides_media_bias_ratings/publication/media_bias_rating,"
            "allsides_media_bias_ratings/publication/source_url,"
            "allsides_media_bias_ratings/publication/allsides_url\n"
            "Good Source,News Media,Center,https://www.example.com/,https://allsides.com/1\n"
            "No Label,News Media,,https://www.nolabel.com/,https://allsides.com/2\n",
            encoding="utf-8",
        )
        index = _build_allsides_index(csv_path)
        assert "example.com" in index
        assert "nolabel.com" not in index


# ---------------------------------------------------------------------------
# _lookup
# ---------------------------------------------------------------------------

class TestLookup:
    @pytest.fixture
    def index(self):
        """Small in-memory AllSides domain index used by the lookup tests."""
        return {
            "cnn.com":       ("CNN",         "Lean Left"),
            "foxnews.com":   ("Fox News",    "Right"),
            "reuters.com":   ("Reuters",     "Center"),
            "bbc.co.uk":     ("BBC",         "Center"),
            "nytimes.com":   ("NY Times",    "Lean Left"),
        }

    def test_exact_match(self, index):
        name, label, matched = _lookup("cnn.com", index)
        assert matched is True
        assert label == "Lean Left"
        assert name == "CNN"

    def test_exact_match_with_www(self, index):
        name, label, matched = _lookup("www.foxnews.com", index)
        assert matched is True
        assert label == "Right"

    def test_subdomain_match(self, index):
        # 'opinion.reuters.com' ends with '.reuters.com' - should match
        name, label, matched = _lookup("opinion.reuters.com", index)
        assert matched is True
        assert label == "Center"

    def test_no_match_returns_no_label(self, index):
        name, label, matched = _lookup("unknown-outlet-xyz123.io", index)
        assert matched is False
        assert label == "no_label"
        assert name == ""

    def test_empty_source_name(self, index):
        name, label, matched = _lookup("", index)
        assert matched is False
        assert label == "no_label"

    def test_case_insensitive(self, index):
        name, label, matched = _lookup("CNN.COM", index)
        assert matched is True

    def test_full_url_as_source_name(self, index):
        # Some articles might store full URLs instead of bare domains
        name, label, matched = _lookup("https://www.reuters.com/news/article", index)
        assert matched is True
        assert label == "Center"


# ---------------------------------------------------------------------------
# run() - full pipeline with tmp dirs
# ---------------------------------------------------------------------------

class TestRunPipeline:

    def _make_allsides_csv(self, path: Path) -> Path:
        csv_path = path / "AllSides_Rating.csv"
        csv_path.write_text(
            "allsides_media_bias_ratings/publication/source_name,"
            "allsides_media_bias_ratings/publication/source_type,"
            "allsides_media_bias_ratings/publication/media_bias_rating,"
            "allsides_media_bias_ratings/publication/source_url,"
            "allsides_media_bias_ratings/publication/allsides_url\n"
            "CNN,News Media,Lean Left,https://www.cnn.com/,https://allsides.com/cnn\n"
            "Fox News,News Media,Right,https://www.foxnews.com/,https://allsides.com/fox\n"
            "Reuters,News Media,Center,https://www.reuters.com/,https://allsides.com/reuters\n",
            encoding="utf-8",
        )
        return csv_path

    def _make_articles(self, path: Path) -> list:
        articles = [
            {"record_id": "a001", "source_name": "cnn.com",       "title": "CNN article"},
            {"record_id": "a002", "source_name": "foxnews.com",    "title": "Fox article"},
            {"record_id": "a003", "source_name": "reuters.com",    "title": "Reuters article"},
            {"record_id": "a004", "source_name": "unknown.co.nz",  "title": "Unknown article"},
        ]
        for art in articles:
            (path / f"{art['record_id']}.json").write_text(json.dumps(art), encoding="utf-8")
        return articles

    def test_dry_run_writes_no_files(self, tmp_path):
        arts_dir = tmp_path / "articles"
        arts_dir.mkdir()
        csv_path = self._make_allsides_csv(tmp_path)
        self._make_articles(arts_dir)

        out_csv   = tmp_path / "labels.csv"
        out_excel = tmp_path / "ref.xlsx"
        out_dir   = tmp_path / "labeled"

        run(arts_dir, csv_path, out_csv, out_excel, out_dir, dry_run=True)

        assert not out_csv.exists(),   "CSV must not be written in dry-run mode"
        assert not out_excel.exists(), "Excel must not be written in dry-run mode"
        assert not out_dir.exists(),   "labeled_articles/ must not be created in dry-run mode"

    def test_csv_headers_and_row_count(self, tmp_path):
        arts_dir = tmp_path / "articles"
        arts_dir.mkdir()
        csv_path = self._make_allsides_csv(tmp_path)
        self._make_articles(arts_dir)

        out_csv   = tmp_path / "labels.csv"
        out_excel = tmp_path / "ref.xlsx"
        out_dir   = tmp_path / "labeled"

        run(arts_dir, csv_path, out_csv, out_excel, out_dir, dry_run=False)

        with open(out_csv, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert set(rows[0].keys()) == {"record_id", "source_name", "allsides_label", "allsides_source_name", "matched"}
        assert len(rows) == 4

    def test_csv_correct_labels(self, tmp_path):
        arts_dir = tmp_path / "articles"
        arts_dir.mkdir()
        csv_path = self._make_allsides_csv(tmp_path)
        self._make_articles(arts_dir)

        out_csv   = tmp_path / "labels.csv"
        out_excel = tmp_path / "ref.xlsx"
        out_dir   = tmp_path / "labeled"

        run(arts_dir, csv_path, out_csv, out_excel, out_dir, dry_run=False)

        with open(out_csv, newline="", encoding="utf-8") as f:
            by_id = {r["record_id"]: r for r in csv.DictReader(f)}

        assert by_id["a001"]["allsides_label"] == "Lean Left"
        assert by_id["a001"]["matched"] == "True"
        assert by_id["a002"]["allsides_label"] == "Right"
        assert by_id["a003"]["allsides_label"] == "Center"
        assert by_id["a004"]["allsides_label"] == "no_label"
        assert by_id["a004"]["matched"] == "False"

    def test_labeled_folders_created(self, tmp_path):
        arts_dir = tmp_path / "articles"
        arts_dir.mkdir()
        csv_path = self._make_allsides_csv(tmp_path)
        self._make_articles(arts_dir)

        out_dir = tmp_path / "labeled"
        run(arts_dir, csv_path, tmp_path / "l.csv", tmp_path / "r.xlsx", out_dir, dry_run=False)

        # Only folders that have at least one article should exist (all 7 are created regardless)
        assert out_dir.exists()
        for folder in LABEL_FOLDERS.values():
            assert (out_dir / folder).exists(), f"Expected folder {folder} to exist"

    def test_labeled_file_counts_match_labels(self, tmp_path):
        arts_dir = tmp_path / "articles"
        arts_dir.mkdir()
        csv_path = self._make_allsides_csv(tmp_path)
        self._make_articles(arts_dir)

        out_dir = tmp_path / "labeled"
        run(arts_dir, csv_path, tmp_path / "l.csv", tmp_path / "r.xlsx", out_dir, dry_run=False)

        lean_left_files = list((out_dir / "Lean_Left").iterdir())
        right_files     = list((out_dir / "Right").iterdir())
        center_files    = list((out_dir / "Center").iterdir())
        no_label_files  = list((out_dir / "no_label").iterdir())

        assert len(lean_left_files) == 1  # cnn.com = Lean Left
        assert len(right_files)     == 1  # foxnews.com = Right
        assert len(center_files)    == 1  # reuters.com = Center
        assert len(no_label_files)  == 1  # unknown.co.nz = no_label

    def test_total_copied_files_equals_input_count(self, tmp_path):
        arts_dir = tmp_path / "articles"
        arts_dir.mkdir()
        csv_path = self._make_allsides_csv(tmp_path)
        self._make_articles(arts_dir)

        out_dir = tmp_path / "labeled"
        run(arts_dir, csv_path, tmp_path / "l.csv", tmp_path / "r.xlsx", out_dir, dry_run=False)

        total_copied = sum(len(list(d.iterdir())) for d in out_dir.iterdir() if d.is_dir())
        assert total_copied == 4

    def test_missing_articles_dir_raises(self, tmp_path):
        csv_path = self._make_allsides_csv(tmp_path)
        with pytest.raises(FileNotFoundError, match="Articles directory not found"):
            run(
                tmp_path / "nonexistent_articles",
                csv_path,
                tmp_path / "l.csv",
                tmp_path / "r.xlsx",
                tmp_path / "labeled",
                dry_run=False,
            )

    def test_missing_allsides_csv_raises(self, tmp_path):
        arts_dir = tmp_path / "articles"
        arts_dir.mkdir()
        self._make_articles(arts_dir)
        with pytest.raises(FileNotFoundError, match="AllSides CSV not found"):
            run(
                arts_dir,
                tmp_path / "nonexistent.csv",
                tmp_path / "l.csv",
                tmp_path / "r.xlsx",
                tmp_path / "labeled",
                dry_run=False,
            )

    def test_article_with_no_source_name_goes_to_no_label(self, tmp_path):
        arts_dir = tmp_path / "articles"
        arts_dir.mkdir()
        csv_path = self._make_allsides_csv(tmp_path)
        (arts_dir / "nosrc.json").write_text(
            json.dumps({"record_id": "nosrc-001", "title": "No source"}),
            encoding="utf-8",
        )

        out_dir = tmp_path / "labeled"
        run(arts_dir, csv_path, tmp_path / "l.csv", tmp_path / "r.xlsx", out_dir, dry_run=False)

        with open(tmp_path / "l.csv", newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert rows[0]["allsides_label"] == "no_label"
        assert rows[0]["matched"] == "False"

    def test_corrupt_json_file_is_skipped_with_warning(self, tmp_path, caplog):
        import logging
        arts_dir = tmp_path / "articles"
        arts_dir.mkdir()
        csv_path = self._make_allsides_csv(tmp_path)
        (arts_dir / "good.json").write_text(
            json.dumps({"record_id": "good-001", "source_name": "cnn.com"}),
            encoding="utf-8",
        )
        (arts_dir / "corrupt.json").write_text("{not valid json {{", encoding="utf-8")

        out_dir = tmp_path / "labeled"
        with caplog.at_level(logging.WARNING):
            run(arts_dir, csv_path, tmp_path / "l.csv", tmp_path / "r.xlsx", out_dir, dry_run=False)

        assert any("corrupt.json" in r.message for r in caplog.records if r.levelno == logging.WARNING), \
            "Expected a WARNING log mentioning the corrupt file"

        # Good file was still processed
        with open(tmp_path / "l.csv", newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 1
        assert rows[0]["record_id"] == "good-001"

    def test_excel_two_sheets(self, tmp_path):
        pytest.importorskip("openpyxl")
        import openpyxl

        arts_dir = tmp_path / "articles"
        arts_dir.mkdir()
        csv_path = self._make_allsides_csv(tmp_path)
        self._make_articles(arts_dir)

        out_excel = tmp_path / "ref.xlsx"
        run(arts_dir, csv_path, tmp_path / "l.csv", out_excel, tmp_path / "labeled", dry_run=False)

        wb = openpyxl.load_workbook(out_excel)
        assert wb.sheetnames == ["Outlet Labels", "Match Summary"]

    def test_excel_outlet_labels_headers(self, tmp_path):
        pytest.importorskip("openpyxl")
        import openpyxl

        arts_dir = tmp_path / "articles"
        arts_dir.mkdir()
        csv_path = self._make_allsides_csv(tmp_path)
        self._make_articles(arts_dir)

        out_excel = tmp_path / "ref.xlsx"
        run(arts_dir, csv_path, tmp_path / "l.csv", out_excel, tmp_path / "labeled", dry_run=False)

        wb = openpyxl.load_workbook(out_excel)
        ws = wb["Outlet Labels"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert headers == ["source_name", "allsides_source_name", "allsides_label", "matched", "article_count"]

    def test_excel_match_summary_totals(self, tmp_path):
        pytest.importorskip("openpyxl")
        import openpyxl

        arts_dir = tmp_path / "articles"
        arts_dir.mkdir()
        csv_path = self._make_allsides_csv(tmp_path)
        self._make_articles(arts_dir)

        out_excel = tmp_path / "ref.xlsx"
        run(arts_dir, csv_path, tmp_path / "l.csv", out_excel, tmp_path / "labeled", dry_run=False)

        wb = openpyxl.load_workbook(out_excel)
        ws = wb["Match Summary"]
        summary = {row[0]: row[1] for row in ws.iter_rows(values_only=True) if row[0]}
        assert summary["Total articles"] == 4
        assert summary["Matched"] == 3
        assert summary["Unmatched"] == 1
