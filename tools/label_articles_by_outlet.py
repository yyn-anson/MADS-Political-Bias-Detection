"""
Label 473,989 article JSONs by media outlet bias using AllSides ground-truth ratings.

Usage
-----
    # Full run: write CSV, Excel, and organize article files
    python tools/label_articles_by_outlet.py

    # Dry run: print statistics without writing any files
    python tools/label_articles_by_outlet.py --dry-run

    # Custom paths
    python tools/label_articles_by_outlet.py \\
        --articles  data/articles \\
        --allsides  data/allsides/AllSides_Rating.csv \\
        --out-csv   data/article_outlet_labels.csv \\
        --out-excel data/outlet_bias_reference.xlsx \\
        --out-dir   data/labeled_articles

Output files
------------
data/article_outlet_labels.csv
    record_id, source_name, allsides_label, allsides_source_name, matched

data/outlet_bias_reference.xlsx
    Sheet "Outlet Labels"  — unique outlet -> AllSides label + match confidence
    Sheet "Match Summary"  — per-label counts and match-rate statistics

data/labeled_articles/
    Left/ Lean_Left/ Center/ Lean_Right/ Right/ Mixed/ no_label/
    Each article JSON is *copied* (not moved) into the appropriate subfolder.
"""

import argparse
import csv
import json
import logging
import re
import shutil
import sys
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import urlparse

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── label folders (spaces replaced with underscores for filesystem safety) ──
LABEL_FOLDERS = {
    "Left":       "Left",
    "Lean Left":  "Lean_Left",
    "Center":     "Center",
    "Lean Right": "Lean_Right",
    "Right":      "Right",
    "Mixed":      "Mixed",
    "no_label":   "no_label",
}


# ---------------------------------------------------------------------------
# Domain normalisation
# ---------------------------------------------------------------------------

def _normalise_domain(raw: str) -> str:
    """Strip protocol, www., trailing slash, lowercase.

    Handles both bare domains ('cnn.com') and full URLs
    ('https://www.cnn.com/').
    """
    raw = raw.strip().lower()
    # Add a fake scheme if none present so urlparse works reliably
    if not raw.startswith(("http://", "https://")):
        raw = "https://" + raw
    host = urlparse(raw).netloc or urlparse(raw).path
    # Remove www. prefix
    host = re.sub(r"^www\.", "", host)
    # Remove port if present
    host = host.split(":")[0]
    return host


def _build_allsides_index(csv_path: Path) -> dict:
    """Return {normalised_domain: (source_name, media_bias_rating)}."""
    index = {}
    with open(csv_path, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            url   = row.get("allsides_media_bias_ratings/publication/source_url", "").strip()
            name  = row.get("allsides_media_bias_ratings/publication/source_name", "").strip()
            label = row.get("allsides_media_bias_ratings/publication/media_bias_rating", "").strip()
            if not url or not label:
                continue
            domain = _normalise_domain(url)
            if domain:
                index[domain] = (name, label)
    logger.info("AllSides index built: %d outlets", len(index))
    return index


def _lookup(source_name: str, index: dict) -> tuple:
    """Return (allsides_source_name, label, matched).

    Tries exact match first, then checks if source_name is a suffix of any
    AllSides domain (catches subdomains like 'politics.example.com').
    """
    norm = _normalise_domain(source_name)

    # Exact match
    if norm in index:
        name, label = index[norm]
        return name, label, True

    # Suffix / subdomain match  (e.g. 'abcnews.go.com' matches 'go.com'... skip too-short)
    for domain, (name, label) in index.items():
        if len(domain) > 4 and norm.endswith("." + domain):
            return name, label, True

    # Prefix match — AllSides domain contained in article source_name
    for domain, (name, label) in index.items():
        if len(domain) > 4 and domain in norm:
            return name, label, True

    return "", "no_label", False


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(
    articles_dir: Path,
    allsides_csv: Path,
    out_csv: Path,
    out_excel: Path,
    out_labeled: Path,
    dry_run: bool,
) -> None:
    if not articles_dir.exists():
        raise FileNotFoundError(f"Articles directory not found: {articles_dir}")
    if not allsides_csv.exists():
        raise FileNotFoundError(f"AllSides CSV not found: {allsides_csv}")

    index = _build_allsides_index(allsides_csv)

    article_files = list(articles_dir.glob("*.json"))
    logger.info("Found %d article JSON files", len(article_files))

    # ── first pass: collect all rows ────────────────────────────────────────
    rows = []           # (record_id, source_name, allsides_label, allsides_source_name, matched)
    outlet_labels = {}  # source_name -> (allsides_source_name, label, matched)
    label_counter = Counter()

    for i, fpath in enumerate(article_files):
        if i % 50_000 == 0 and i > 0:
            logger.info("  processed %d / %d ...", i, len(article_files))
        try:
            with open(fpath, encoding="utf-8") as fh:
                article = json.load(fh)
        except Exception as exc:
            logger.warning("Could not read %s: %s", fpath.name, exc)
            continue

        record_id = article.get("record_id")
        if record_id is None:
            logger.warning("Article %s has no 'record_id' field; using filename as fallback.", fpath.name)
            record_id = fpath.stem

        source_name = article.get("source_name", "").strip()
        if not source_name:
            logger.warning("Article %s has no 'source_name'; it will be labeled 'no_label'.", fpath.name)

        if source_name not in outlet_labels:
            outlet_labels[source_name] = _lookup(source_name, index)

        allsides_source, label, matched = outlet_labels[source_name]
        label_counter[label] += 1
        rows.append((record_id, source_name, label, allsides_source, matched, fpath))

    total      = len(rows)
    matched_n  = sum(1 for r in rows if r[4])
    unmatched_n = total - matched_n

    # ── summary ─────────────────────────────────────────────────────────────
    logger.info("─" * 60)
    logger.info("Total articles  : %d", total)
    logger.info("Matched         : %d  (%.1f%%)", matched_n, 100 * matched_n / max(total, 1))
    logger.info("Unmatched       : %d  (%.1f%%)", unmatched_n, 100 * unmatched_n / max(total, 1))
    logger.info("Per-label breakdown:")
    for lbl in ["Left", "Lean Left", "Center", "Lean Right", "Right", "Mixed", "no_label"]:
        logger.info("  %-12s : %d", lbl, label_counter.get(lbl, 0))
    logger.info("─" * 60)

    if dry_run:
        logger.info("Dry run — no files written.")
        return

    # ── write CSV ────────────────────────────────────────────────────────────
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with open(out_csv, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["record_id", "source_name", "allsides_label", "allsides_source_name", "matched"])
        for record_id, source_name, label, allsides_source, matched, _ in rows:
            writer.writerow([record_id, source_name, label, allsides_source, matched])
    logger.info("CSV written: %s", out_csv)

    # ── write Excel ──────────────────────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        logger.warning("openpyxl not installed — skipping Excel output.  Install with: pip install openpyxl")
        out_excel = None

    if out_excel is not None:
        wb = openpyxl.Workbook()

        # Sheet 1: Outlet Labels
        ws1 = wb.active
        ws1.title = "Outlet Labels"
        header = ["source_name", "allsides_source_name", "allsides_label", "matched", "article_count"]
        ws1.append(header)
        bold = Font(bold=True)
        for cell in ws1[1]:
            cell.font = bold

        outlet_counts = Counter(r[1] for r in rows)
        outlet_rows = sorted(
            outlet_labels.items(),
            key=lambda kv: outlet_counts.get(kv[0], 0),
            reverse=True,
        )
        for src, (allsides_src, lbl, matched) in outlet_rows:
            ws1.append([src, allsides_src, lbl, matched, outlet_counts.get(src, 0)])

        # Sheet 2: Match Summary
        ws2 = wb.create_sheet("Match Summary")
        ws2.append(["Metric", "Value"])
        for cell in ws2[1]:
            cell.font = bold
        ws2.append(["Total articles", total])
        ws2.append(["Matched", matched_n])
        ws2.append(["Unmatched", unmatched_n])
        ws2.append(["Match rate (%)", round(100 * matched_n / max(total, 1), 2)])
        ws2.append(["Unique outlets", len(outlet_labels)])
        ws2.append(["Matched outlets", sum(1 for _, (_, _, m) in outlet_labels.items() if m)])
        ws2.append([])
        ws2.append(["Label", "Article count", "Percentage"])
        for lbl in ["Left", "Lean Left", "Center", "Lean Right", "Right", "Mixed", "no_label"]:
            cnt = label_counter.get(lbl, 0)
            ws2.append([lbl, cnt, round(100 * cnt / max(total, 1), 2)])

        out_excel.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_excel)
        logger.info("Excel written: %s", out_excel)

    # ── copy articles into labeled subfolders ────────────────────────────────
    out_labeled.mkdir(parents=True, exist_ok=True)
    for folder in LABEL_FOLDERS.values():
        (out_labeled / folder).mkdir(exist_ok=True)

    copied = 0
    for record_id, source_name, label, allsides_source, matched, fpath in rows:
        folder_name = LABEL_FOLDERS.get(label, "no_label")
        dest = out_labeled / folder_name / fpath.name
        try:
            shutil.copy2(fpath, dest)
            copied += 1
        except Exception as exc:
            logger.warning("Could not copy %s: %s", fpath.name, exc)

    logger.info("Copied %d article files into %s", copied, out_labeled)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_args() -> argparse.Namespace:
    repo_root = Path(__file__).resolve().parent.parent

    parser = argparse.ArgumentParser(
        description="Label article JSONs by AllSides outlet bias rating.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--articles",
        type=Path,
        default=repo_root / "data" / "articles",
        help="Directory containing article JSON files.",
    )
    parser.add_argument(
        "--allsides",
        type=Path,
        default=repo_root / "data" / "allsides" / "AllSides_Rating.csv",
        help="Path to AllSides_Rating.csv.",
    )
    parser.add_argument(
        "--out-csv",
        type=Path,
        default=repo_root / "data" / "article_outlet_labels.csv",
        help="Output path for the per-article label CSV.",
    )
    parser.add_argument(
        "--out-excel",
        type=Path,
        default=repo_root / "data" / "outlet_bias_reference.xlsx",
        help="Output path for the outlet reference Excel workbook.",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=repo_root / "data" / "labeled_articles",
        help="Root directory for label subfolders.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print statistics without writing any files.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    run(
        articles_dir=args.articles,
        allsides_csv=args.allsides,
        out_csv=args.out_csv,
        out_excel=args.out_excel,
        out_labeled=args.out_dir,
        dry_run=args.dry_run,
    )
